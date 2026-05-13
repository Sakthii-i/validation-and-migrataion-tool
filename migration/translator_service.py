import csv
import json
import os
import re
import threading
import time
from typing import Any, Callable, Dict, List, Optional, Tuple

import httpx
import sqlglot

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

try:
    import google.generativeai as genai
except Exception:
    genai = None

try:
    from anthropic import Anthropic
except Exception:
    Anthropic = None

from .ast_transformer import BigQueryToDatabricksTransformer, ExpressionOptimizer
from .rule_engine import RuleEngine
from .sql_processor import ExpressionCache, QueryChunker, SQLPreprocessor
from .translation_cache import TranslationCache
from .complexity_analyzer import QueryComplexityAnalyzer
from .validator import LLMFixerPrompt, SQLValidator

BASE_DIR = os.path.dirname(__file__)

PROVIDER_MODEL_OPTIONS: Dict[str, List[str]] = {
    "OpenAI": ["gpt-5-nano", "gpt-5-mini", "gpt-4.1-mini", "gpt-4.1"],
    "Gemini": ["gemini-2.5-flash", "gemini-2.5-pro"],
    "Claude": ["claude-3-5-haiku-latest", "claude-3-5-sonnet-latest", "claude-3-opus-latest"],
}


def load_conversion_rules() -> Tuple[List[Dict[str, str]], Dict[str, List[Dict[str, str]]], List[Dict[str, str]]]:
    """Load conversion rules from xlsx primary source with csv fallback."""
    rules_dict: Dict[str, List[Dict[str, str]]] = {}
    rules_list: List[Dict[str, str]] = []
    edge_cases: List[Dict[str, str]] = []

    xlsx_path = os.path.join(BASE_DIR, "learnings from ciq.xlsx")
    try:
        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        if "functions" in wb.sheetnames:
            ws = wb["functions"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h).strip() if h else "" for h in rows[0]]
                for row in rows[1:]:
                    row_dict = {
                        headers[i]: (str(row[i]).strip() if row[i] is not None else "")
                        for i in range(len(headers))
                    }
                    if row_dict.get("bigquery_syntax") in ("", "-", "null", "None"):
                        continue
                    rules_list.append(row_dict)
                    cat = row_dict.get("category", "")
                    if cat not in rules_dict:
                        rules_dict[cat] = []
                    if row_dict.get("bigquery_syntax") and row_dict.get("databricks_sql_syntax"):
                        rules_dict[cat].append(
                            {
                                "bq": row_dict["bigquery_syntax"],
                                "db": row_dict["databricks_sql_syntax"],
                                "example_bq": row_dict.get("example_bq", ""),
                                "example_dbsql": row_dict.get("example_dbsql", ""),
                            }
                        )

        if "edge cases" in wb.sheetnames:
            ws = wb["edge cases"]
            for row in ws.iter_rows(values_only=True):
                vals = [str(v).strip() if v else "" for v in row]
                scenario = vals[0] if len(vals) > 0 else ""
                bq_syntax = vals[2] if len(vals) > 2 else ""
                dbx_syntax = vals[4] if len(vals) > 4 else ""
                note = vals[6] if len(vals) > 6 else ""
                if bq_syntax and dbx_syntax and scenario:
                    edge_cases.append(
                        {
                            "scenario": scenario,
                            "bq": bq_syntax,
                            "dbx": dbx_syntax,
                            "note": note,
                        }
                    )

        if "data type" in wb.sheetnames:
            ws = wb["data type"]
            for row in ws.iter_rows(values_only=True, min_row=2):
                vals = [str(v).strip().replace("\xa0", "") if v else "" for v in row]
                bq_type = vals[0] if len(vals) > 0 else ""
                dbx_type = vals[1] if len(vals) > 1 else ""
                if bq_type and dbx_type and bq_type != dbx_type:
                    cat = "Data Types (CIQ)"
                    if cat not in rules_dict:
                        rules_dict[cat] = []
                    rules_dict[cat].append({"bq": bq_type, "db": dbx_type, "example_bq": "", "example_dbsql": ""})

        wb.close()

    except FileNotFoundError:
        try:
            with open(os.path.join(BASE_DIR, "conversion_rules_bigquery.csv"), encoding="utf-8") as f:
                for row in csv.DictReader(f):
                    rules_list.append(dict(row))
                    cat = row.get("category", "")
                    if cat not in rules_dict:
                        rules_dict[cat] = []
                    if row.get("bigquery_syntax") and row.get("databricks_sql_syntax"):
                        rules_dict[cat].append(
                            {
                                "bq": row["bigquery_syntax"],
                                "db": row["databricks_sql_syntax"],
                                "example_bq": row.get("example_bq", ""),
                                "example_dbsql": row.get("example_dbsql", ""),
                            }
                        )
        except Exception as exc:
            raise RuntimeError(f"Error loading conversion rules: {exc}") from exc
    except Exception as exc:
        raise RuntimeError(f"Error loading Excel rules: {exc}") from exc

    return rules_list, rules_dict, edge_cases


class TranslatorService:
    """Service wrapper for the existing translation pipeline."""

    def __init__(self):
        self._lock = threading.Lock()
        self._components: Optional[Dict[str, Any]] = None
        # In-memory aggregate statistics for the current process.
        # These are intentionally simple and process-local; if we need
        # per-user or per-session isolation later, this can be extended.
        self._session_stats: Dict[str, Any] = {
            "total_queries_processed": 0,
            "successful_migrations": 0,
            "failed_migrations": 0,
            "simple_queries": 0,
            "medium_queries": 0,
            "complex_queries": 0,
            "low_risk_migrations": 0,
            "medium_risk_migrations": 0,
            "high_risk_migrations": 0,
            "average_complexity_score": 0.0,
        }

    def _init_components(self) -> Dict[str, Any]:
        rules_list, _, edge_cases = load_conversion_rules()
        cache_db_path = os.environ.get("TRANSLATION_CACHE_DB_PATH", os.path.join(BASE_DIR, "translation_cache.db"))
        return {
            "preprocessor": SQLPreprocessor(),
            "chunker": QueryChunker(max_chunk_size=500),
            "transformer": BigQueryToDatabricksTransformer(),
            "rule_engine": RuleEngine(rules_list, edge_cases),
            "cache": TranslationCache(db_path=cache_db_path),
            "validator": SQLValidator(),
            "expr_cache": ExpressionCache(),
        }

    @property
    def components(self) -> Dict[str, Any]:
        if self._components is None:
            with self._lock:
                if self._components is None:
                    self._components = self._init_components()
        return self._components

    def _update_session_stats(self, complexity: Dict[str, Any], is_success: bool) -> Dict[str, Any]:
        """
        Update process-wide aggregate statistics from a single query's
        complexity and outcome. This is a lightweight approximation of
        Databricks Labs DQX-style rollups for the current runtime.
        """
        stats = self._session_stats

        stats["total_queries_processed"] += 1
        if is_success:
            stats["successful_migrations"] += 1
        else:
            stats["failed_migrations"] += 1

        level = (complexity.get("complexity_level") or "").upper()
        if level == "SIMPLE":
            stats["simple_queries"] += 1
        elif level == "MEDIUM":
            stats["medium_queries"] += 1
        elif level == "COMPLEX":
            stats["complex_queries"] += 1

        risk = (complexity.get("estimated_conversion_risk") or "").lower()
        if risk == "low":
            stats["low_risk_migrations"] += 1
        elif risk == "medium":
            stats["medium_risk_migrations"] += 1
        elif risk == "high":
            stats["high_risk_migrations"] += 1

        score = int(complexity.get("complexity_score") or 0)
        n = stats["total_queries_processed"]
        prev_avg = float(stats.get("average_complexity_score") or 0.0)
        stats["average_complexity_score"] = prev_avg + (score - prev_avg) / float(max(n, 1))

        return stats.copy()

    @staticmethod
    def _resolve_api_key(explicit_key: Optional[str], env_key: str) -> str:
        api_key = (explicit_key or "").strip() or os.environ.get(env_key, "").strip()
        return api_key

    @staticmethod
    def _transpile_to_databricks(sql_text: str, source_dialect: str = "bigquery") -> Tuple[str, Optional[str]]:
        """Deterministic sqlglot transpile: source SQL AST -> Databricks SQL."""
        dialect = "snowflake" if (source_dialect or "").strip().lower() == "snowflake" else "bigquery"
        try:
            trees = [t for t in sqlglot.parse(sql_text, read=dialect) if t is not None]
            if not trees:
                raise sqlglot.errors.ParseError("Empty parse result")
            transpiled = ";\n".join(tree.sql(dialect="databricks", pretty=True) for tree in trees)
            # Safety check: if transpiled output lost >40% of content,
            # the AST parser silently dropped syntax it couldn't handle.
            orig_len = len(sql_text.strip())
            trans_len = len(transpiled.strip())
            if orig_len > 2000 and trans_len < orig_len * 0.6:
                return sql_text, (
                    f"sqlglot transpile lost significant content "
                    f"({orig_len} -> {trans_len} chars, "
                    f"{(1 - trans_len / orig_len) * 100:.0f}% loss)"
                )
            return transpiled, None
        except Exception as exc:
            return sql_text, f"sqlglot {dialect} parse/transpile failed: {exc}"

    @staticmethod
    def _validate_databricks_parse(sql_text: str) -> Optional[str]:
        """Returns None if parseable as Databricks SQL, else error message."""
        try:
            trees = sqlglot.parse(sql_text, read="databricks")
            if not any(t is not None for t in trees):
                raise sqlglot.errors.ParseError("Empty parse result")
            return None
        except Exception as exc:
            return f"Databricks re-parse failed: {exc}"

    @staticmethod
    def _normalize_databricks_host(host: str) -> str:
        normalized = (host or "").strip()
        if not normalized:
            raise ValueError("Databricks host is required")
        if not normalized.startswith("http://") and not normalized.startswith("https://"):
            normalized = f"https://{normalized}"
        return normalized.rstrip("/")

    @staticmethod
    def _normalize_warehouse_id(warehouse_value: str) -> str:
        raw = (warehouse_value or "").strip()
        if not raw:
            return ""

        match = re.search(r"/sql/1\.0/warehouses/([A-Za-z0-9_-]+)", raw)
        if match:
            return match.group(1)

        match = re.search(r"warehouses/([A-Za-z0-9_-]+)", raw)
        if match:
            return match.group(1)

        return raw

    @staticmethod
    def _extract_rows(
        statement_resp: Dict[str, Any],
        max_rows: int,
        http_client: Optional[httpx.Client] = None,
        headers: Optional[Dict[str, str]] = None,
    ) -> Dict[str, Any]:
        manifest = statement_resp.get("manifest", {}) or {}
        schema = (manifest.get("schema", {}) or {}).get("columns", []) or []
        columns = [str(col.get("name", "")) for col in schema]

        result = statement_resp.get("result", {}) or {}
        data_array = result.get("data_array", []) or []

        if not data_array:
            external_links = result.get("external_links", []) or []
            if external_links and http_client is not None:
                link_info = external_links[0] or {}
                link_url = link_info.get("external_link") or link_info.get("url")
                if link_url:
                    link_resp = http_client.get(link_url, headers=headers)
                    link_resp.raise_for_status()
                    try:
                        link_payload = link_resp.json()
                    except Exception:
                        link_payload = json.loads(link_resp.text)
                    if isinstance(link_payload, dict):
                        data_array = link_payload.get("data_array", []) or link_payload.get("rows", []) or []
                    elif isinstance(link_payload, list):
                        data_array = link_payload

        sample_limit = min(max_rows, 10)
        trimmed_rows = data_array[:sample_limit]
        rows = [
            {columns[i] if i < len(columns) else f"col_{i}": value for i, value in enumerate(row)}
            for row in trimmed_rows
        ]

        total_rows = manifest.get("total_row_count")
        if total_rows is None:
            total_rows = len(data_array)

        return {
            "columns": columns,
            "rows": rows,
            "row_count": int(total_rows) if isinstance(total_rows, (int, float, str)) and str(total_rows).isdigit() else len(data_array),
            "truncated": len(data_array) > sample_limit or (isinstance(total_rows, int) and total_rows > sample_limit),
        }

    def execute_databricks_sql(self, sql: str, databricks_cfg: Dict[str, Any]) -> Dict[str, Any]:
        host = self._normalize_databricks_host(databricks_cfg.get("host", ""))
        token = (databricks_cfg.get("token") or "").strip()
        warehouse_id = self._normalize_warehouse_id(databricks_cfg.get("warehouse_id", ""))

        if not token:
            raise ValueError("Databricks token is required")
        if not warehouse_id:
            raise ValueError("Databricks warehouse_id is required")

        timeout_seconds = int(databricks_cfg.get("timeout_seconds", 90) or 90)
        max_rows = int(databricks_cfg.get("max_rows", 200) or 200)
        catalog = (databricks_cfg.get("catalog") or "").strip() or None
        schema = (databricks_cfg.get("schema") or "").strip() or None

        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        }

        payload: Dict[str, Any] = {
            "statement": sql,
            "warehouse_id": warehouse_id,
            "wait_timeout": "10s",
            "on_wait_timeout": "CONTINUE",
            "disposition": "EXTERNAL_LINKS",
            "format": "JSON_ARRAY",
        }
        if catalog:
            payload["catalog"] = catalog
        if schema:
            payload["schema"] = schema

        with httpx.Client(timeout=timeout_seconds + 15) as client:
            response = client.post(f"{host}/api/2.0/sql/statements", headers=headers, json=payload)
            response.raise_for_status()
            statement_data = response.json()

            statement_id = statement_data.get("statement_id")
            if not statement_id:
                return {
                    "status": (statement_data.get("status") or {}).get("state", "PENDING"),
                    "statement_id": None,
                    "raw": statement_data,
                }

            start_time = time.time()
            while True:
                status = (statement_data.get("status") or {}).get("state", "")
                if status in ("SUCCEEDED", "FAILED", "CANCELED", "CLOSED"):
                    break

                if (time.time() - start_time) > timeout_seconds:
                    raise TimeoutError(f"Timed out waiting for Databricks statement {statement_id}")

                poll_url = f"{host}/api/2.0/sql/statements/{statement_id}"
                poll_resp = client.get(poll_url, headers=headers)
                poll_resp.raise_for_status()
                statement_data = poll_resp.json()
                time.sleep(1)

            final_status = (statement_data.get("status") or {}).get("state", "")
            if final_status != "SUCCEEDED":
                err = (statement_data.get("status") or {}).get("error") or {}
                message = err.get("message") or f"Databricks statement ended with status: {final_status}"
                return {
                    "status": final_status,
                    "statement_id": statement_id,
                    "error": message,
                }

            extracted = self._extract_rows(statement_data, max_rows=max_rows, http_client=client, headers=headers)
            return {
                "status": final_status,
                "statement_id": statement_id,
                **extracted,
            }

    

    def get_llm_client(self, provider: str, api_key: Optional[str]) -> Optional[Any]:
        provider_norm = (provider or "OpenAI").strip().lower()

        if provider_norm == "openai":
            if OpenAI is None:
                return None
            key = self._resolve_api_key(api_key, "OPENAI_API_KEY")
            return OpenAI(api_key=key) if key else None

        if provider_norm == "gemini":
            if genai is None:
                return None
            key = self._resolve_api_key(api_key, "GEMINI_API_KEY")
            if not key:
                return None
            genai.configure(api_key=key)
            return genai

        if provider_norm == "claude":
            if Anthropic is None:
                return None
            key = self._resolve_api_key(api_key, "ANTHROPIC_API_KEY")
            return Anthropic(api_key=key) if key else None

        return None

    def run_pipeline(
        self,
        bq_sql: str,
        model: str,
        provider: str = "OpenAI",
        api_key: Optional[str] = None,
        source_engine: str = "bigquery",
        force_llm: bool = False,
        use_llm: bool = True,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> Tuple[str, str, Dict[str, Any], Optional[str]]:
        components = self.components
        client = self.get_llm_client(provider, api_key) if use_llm else None

        stats: Dict[str, Any] = {
            "steps": [],
            "llm_calls": 0,
            "cache_hits": 0,
            "chunks": 0,
            "errors": [],
        }

        def _repair_common_llm_mistakes(sql_text: str) -> str:
            # Fix hallucinated DATE_TRUNC('date', 'WEEK') style output.
            # Databricks requires DATE_TRUNC('UNIT', expr), where expr is not a quoted column name.
            sql_text = re.sub(
                r"\bDATE_TRUNC\s*\(\s*'([A-Za-z_][A-Za-z0-9_\.]*)'\s*,\s*'(DAY|WEEK|MONTH|QUARTER|YEAR|HOUR|MINUTE|SECOND)'\s*\)",
                r"DATE_TRUNC('\2', \1)",
                sql_text,
                flags=re.IGNORECASE,
            )
            return sql_text

        def _post_llm_cleanup(sql_text: str) -> str:
            sql_text = _repair_common_llm_mistakes(sql_text)
            cleaned = components["rule_engine"].apply_rules(sql_text)
            cleaned = components["rule_engine"].apply_function_translation(cleaned)
            cleaned = ExpressionOptimizer.optimize(cleaned)
            cleaned = _repair_common_llm_mistakes(cleaned)
            return cleaned

        def _is_suspiciously_short_llm_output(original_sql: str, llm_sql: str) -> bool:
            """Reject likely truncated LLM rewrites that would drop large parts of SQL."""
            o = (original_sql or "").strip()
            n = (llm_sql or "").strip()
            if not o or not n:
                return True
            if len(o) < 4000:
                return False
            return len(n) < max(1200, int(len(o) * 0.35))

        cache_version = ":v2026_04_27_full_coverage"
        provider_key = provider.lower().strip() if provider else "openai"
        model_key = re.sub(r"\s+", "-", (model or "").lower())
        cache_suffix = (f":llm:{provider_key}:{model_key}" if use_llm else ":det") + cache_version
        if not force_llm:
            cached = components["cache"].get(bq_sql + cache_suffix)
            if cached:
                stats["cache_hits"] = 1
                stats["steps"].append("Cache hit; skipped all translation steps")
                return cached["translated"], "Served from cache.", stats, None

        sql_no_comments, comment_map = SQLPreprocessor.extract_comments(bq_sql)
        stats["steps"].append(f"Extracted {len(comment_map)} comment(s)")

        sql_no_comments = SQLPreprocessor.convert_dbt_partition_config(sql_no_comments)

        normalized = SQLPreprocessor.clean_sql(sql_no_comments)
        stats["steps"].append(f"Normalized: {len(bq_sql)} -> {len(normalized)} chars")
        llm_context = SQLPreprocessor.remove_comment_placeholders(normalized)

        is_scripting = SQLPreprocessor.is_scripting_block(normalized)
        is_transaction = bool(re.search(r"\bBEGIN\s+TRANSACTION\b", normalized, re.IGNORECASE))
        if is_scripting:
            stats["steps"].append("SQL scripting block detected (Databricks 2026 GA)")
        if is_transaction:
            stats["steps"].append("Multi-statement transaction detected (Databricks 2026)")

        chunks = components["chunker"].chunk_query(normalized)
        stats["chunks"] = len(chunks)
        order = components["chunker"].get_translation_order(chunks)
        chunk_map = {c.id: c for c in chunks}
        stats["steps"].append(f"Split into {len(chunks)} chunk(s)")

        translated_map: Dict[str, str] = {}
        total_chunks = len(order)

        for i, chunk_id in enumerate(order):
            if progress_callback:
                progress_callback(i + 1, total_chunks)

            chunk = chunk_map[chunk_id]
            cache_key = chunk.sql + cache_suffix
            cached_expr = components["expr_cache"].get(cache_key)
            if cached_expr:
                translated_map[chunk_id] = cached_expr
                continue

            t, transpile_err = self._transpile_to_databricks(chunk.sql, source_engine)
            if transpile_err:
                # Keep deterministic behavior robust for noisy/Jinja-heavy inputs:
                # if direct sqlglot transpile fails, fall back to legacy AST/regex path.
                pre_ast_sql = RuleEngine.apply_pre_ast_translation(chunk.sql)
                t = components["transformer"].transform(pre_ast_sql)
                stats["steps"].append(
                    f"Chunk {chunk_id}: sqlglot transpile failed -> legacy deterministic fallback applied"
                )

            t = components["rule_engine"].apply_rules(t)
            t = components["rule_engine"].apply_function_translation(t)
            t = ExpressionOptimizer.optimize(t)

            # ── Per-chunk: only fix parse errors / validation failures ──
            # Proactive LLM migration is done ONCE on the full assembled
            # query (not per-chunk) to avoid N separate LLM calls.
            parse_err = self._validate_databricks_parse(t)
            v = components["validator"].validate(t)
            needs_llm_fix = False
            llm_prompt = None

            if parse_err and client and not force_llm:
                needs_llm_fix = True
                llm_prompt = LLMFixerPrompt.create_chunk_fix_prompt(t, chunk.sql, parse_err)
                stats["steps"].append(f"Chunk {chunk_id}: Databricks re-parse failed -> sending to {provider}")
            elif parse_err and not client:
                stats["errors"].append(
                    f"Chunk {chunk_id}: {parse_err} (no API key; {provider} fallback unavailable)"
                )
            elif not v.is_valid and client and not force_llm:
                needs_llm_fix = True
                if v.error_type == "RESIDUAL_BQ_SYNTAX":
                    llm_prompt = LLMFixerPrompt.create_residual_fix_prompt(t, chunk.sql, v.error_message or "")
                    stats["steps"].append(f"Chunk {chunk_id}: untranslated BQ patterns detected -> sending to {provider}")
                else:
                    llm_prompt = LLMFixerPrompt.create_chunk_fix_prompt(t, chunk.sql, v.error_message or "")
            elif not v.is_valid and not client:
                stats["errors"].append(
                    f"Chunk {chunk_id}: {v.error_message} (no API key; {provider} fallback unavailable)"
                )

            if needs_llm_fix and llm_prompt and client:
                stats["llm_calls"] += 1
                llm_result, err = self._llm_fix_with_prompt(prompt=llm_prompt, client=client, model=model, provider=provider)
                if err:
                    stats["errors"].append(f"Chunk {chunk_id}: {err}")
                elif llm_result.strip():
                    if _is_suspiciously_short_llm_output(t, llm_result):
                        stats["errors"].append(
                            f"Chunk {chunk_id}: rejected suspiciously short LLM output to prevent SQL loss"
                        )
                    else:
                        t = _post_llm_cleanup(llm_result)
                        stats["steps"].append(f"Chunk {chunk_id}: {provider} fix applied")

            components["expr_cache"].set(cache_key, t)
            translated_map[chunk_id] = t

        source_dialect = "snowflake" if (source_engine or "").strip().lower() == "snowflake" else "bigquery"
        stats["steps"].append(f"Translated {len(order)} chunk(s) via sqlglot (read={source_dialect}, write=databricks); {stats['llm_calls']} LLM call(s)")

        assembled = components["chunker"].reassemble(chunks, translated_map)

        # ── Step 1: Fix parse errors on assembled query ──
        assembled_parse_err = self._validate_databricks_parse(assembled)
        if assembled_parse_err and client and not force_llm:
            stats["llm_calls"] += 1
            prompt = LLMFixerPrompt.create_fix_prompt(llm_context, assembled, assembled_parse_err)
            llm_result, err = self._llm_fix_with_prompt(prompt, client, model, provider=provider)
            if err:
                stats["errors"].append(err)
            elif llm_result.strip():
                if _is_suspiciously_short_llm_output(assembled, llm_result):
                    stats["errors"].append(
                        "Final SQL: rejected suspiciously short LLM parse-fix output to prevent SQL loss"
                    )
                else:
                    assembled = _post_llm_cleanup(llm_result)
                    stats["steps"].append(f"Final SQL Databricks re-parse failed -> {provider} fix applied")
        elif assembled_parse_err and not client:
            stats["errors"].append(
                f"Final SQL: {assembled_parse_err} (no API key; {provider} fallback unavailable)"
            )

        # ── Step 2: Proactive LLM migration on the FULL assembled query ──
        # Deterministic rules don't cover everything, so we send the full
        # query to the LLM once to catch anything the rules missed.
        # This is ONE call on the full query instead of N calls per chunk.
        if client and use_llm and not force_llm:
            stats["llm_calls"] += 1
            prompt = LLMFixerPrompt.create_proactive_migration_prompt(assembled, llm_context)
            llm_result, err = self._llm_fix_with_prompt(prompt, client, model, provider=provider)
            if err:
                stats["errors"].append(f"Proactive LLM migration failed: {err}")
            elif llm_result.strip():
                if _is_suspiciously_short_llm_output(assembled, llm_result):
                    stats["errors"].append(
                        "Final SQL: rejected suspiciously short proactive LLM output to prevent SQL loss"
                    )
                else:
                    assembled = _post_llm_cleanup(llm_result)
                    stats["steps"].append(f"Proactive {provider} migration applied on full query")

        # ── Step 3: Final validation ──
        final_v = components["validator"].validate(assembled)
        if not final_v.is_valid and client and not force_llm:
            stats["llm_calls"] += 1

            if final_v.error_type == "RESIDUAL_BQ_SYNTAX":
                prompt = LLMFixerPrompt.create_residual_fix_prompt(assembled, llm_context, final_v.error_message or "")
                stats["steps"].append(f"Final SQL has untranslated BQ patterns -> sending to {provider}")
            else:
                prompt = LLMFixerPrompt.create_fix_prompt(llm_context, assembled, final_v.error_message or "")

            llm_result, err = self._llm_fix_with_prompt(prompt, client, model, provider=provider)
            if err:
                stats["errors"].append(err)
            else:
                if _is_suspiciously_short_llm_output(assembled, llm_result):
                    stats["errors"].append(
                        "Final SQL: rejected suspiciously short LLM full-query fix output to prevent SQL loss"
                    )
                    llm_result = assembled
                llm_v = components["validator"].validate(llm_result)
                if llm_v.is_valid:
                    assembled = _post_llm_cleanup(llm_result)
                else:
                    stats["errors"].append(f"LLM full-query fix still invalid: {llm_v.error_message}")
            stats["steps"].append("LLM applied full-query fix")
        elif not final_v.is_valid and not client:
            stats["errors"].append(f"Validation failed: {final_v.error_message} (no API key; {provider} fallback unavailable)")
            stats["steps"].append(f"{final_v.error_message}; provide {provider} API key to auto-fix")
        elif not final_v.is_valid:
            stats["errors"].append(f"Validation failed: {final_v.error_message}")
            stats["steps"].append(f"Validation warning: {final_v.error_message}")
        else:
            stats["steps"].append("Final SQL validated")

        final_sql = SQLPreprocessor.restore_comments(assembled, comment_map)
        stats["steps"].append(f"Restored {len(comment_map)} comment(s)")

        if not stats["errors"]:
            components["cache"].set(bq_sql + cache_suffix, final_sql)
            stats["steps"].append("Cached successful translation")
        else:
            stats["steps"].append("Bypassed cache due to validation/LLM errors")

        # ── Complexity & session rollups (DQX-style) ──
        analyzer = QueryComplexityAnalyzer()
        source_sql = bq_sql
        try:
            # If we later add true Snowflake source text, we can branch
            # on source_engine here. For now, we analyze the provided
            # input SQL, which is BigQuery or Snowflake depending on
            # upstream usage.
            complexity = analyzer.analyze(source_sql)
        except Exception as exc:
            complexity = {
                "complexity_level": "SIMPLE",
                "complexity_score": 0,
                "indicators": [],
                "tables_referenced": 0,
                "joins_count": 0,
                "subqueries_count": 0,
                "aggregations": False,
                "window_functions": False,
                "cte_count": 0,
                "set_operations": False,
                "estimated_conversion_risk": "low",
                "conversion_risk_factors": [f"Complexity analysis failed: {exc}"],
                "vendor_specific_functions": [],
            }

        final_error = "\n".join(stats["errors"]) if stats["errors"] else None
        is_success = not final_error
        session_stats = self._update_session_stats(complexity, is_success=is_success)

        stats["complexity"] = complexity
        stats["session"] = session_stats

        explanation = self._build_explanation(stats)
        return final_sql, explanation, stats, final_error

    def _llm_fix_with_prompt(
        self,
        prompt: str,
        client: Any,
        model: str,
        provider: str = "OpenAI",
        max_tokens: int = 16384,
    ) -> Tuple[str, Optional[str]]:
        """Send pre-built prompt to selected provider with truncation retries."""
        max_ceiling = 128_000
        max_retries = 4
        current_max_tokens = max_tokens
        last_result = ""

        provider_norm = (provider or "OpenAI").strip().lower()

        def _extract_sql(text: str) -> str:
            text = text.strip()
            m = re.search(r"```[a-zA-Z]*\n(.*?)```", text, re.DOTALL)
            if m:
                return m.group(1).strip()
            if text.startswith("```"):
                lines = [line for line in text.split("\n") if not line.strip().startswith("```")]
                return "\n".join(lines).strip()
            return text

        for attempt in range(max_retries):
            try:
                if provider_norm == "openai":
                    resp = client.chat.completions.create(
                        model=model,
                        messages=[{"role": "user", "content": prompt}],
                        max_tokens=current_max_tokens,
                        temperature=0,
                    )
                    parts = []
                    for choice in getattr(resp, "choices", []) or []:
                        msg = getattr(choice, "message", None)
                        text = getattr(msg, "content", None)
                        if text:
                            parts.append(text)
                    result = "\n".join(parts).strip()
                    last_result = result

                    result = _extract_sql(result)

                    stop_reason = ""
                    if getattr(resp, "choices", None):
                        stop_reason = str(getattr(resp.choices[0], "finish_reason", ""))
                    if stop_reason != "length":
                        return result, None

                elif provider_norm == "gemini":
                    model_client = client.GenerativeModel(
                        model_name=model,
                        system_instruction="You are an expert SQL transpiler. Output ONLY valid SQL. No markdown formatting (no ``` fences), no text explanations."
                    )
                    resp = model_client.generate_content(
                        prompt,
                        generation_config=client.types.GenerationConfig(
                            temperature=0,
                            max_output_tokens=current_max_tokens,
                        ),
                    )
                    result = (getattr(resp, "text", "") or "").strip()
                    if not result and getattr(resp, "candidates", None):
                        parts = []
                        for part in getattr(resp.candidates[0].content, "parts", []) or []:
                            text = getattr(part, "text", None)
                            if text:
                                parts.append(text)
                        result = "\n".join(parts).strip()
                    last_result = result

                    result = _extract_sql(result)

                    stop_reason = ""
                    if getattr(resp, "candidates", None):
                        stop_reason = str(getattr(resp.candidates[0], "finish_reason", ""))
                    if "MAX_TOKENS" not in stop_reason:
                        return result, None

                elif provider_norm == "claude":
                    resp = client.messages.create(
                        model=model,
                        max_tokens=current_max_tokens,
                        temperature=0,
                        system="You are an expert SQL transpiler. Output ONLY valid SQL. No markdown formatting (no ``` fences), no text explanations.",
                        messages=[{"role": "user", "content": prompt}],
                    )
                    parts = []
                    for block in getattr(resp, "content", []) or []:
                        text = getattr(block, "text", None)
                        if text:
                            parts.append(text)
                    result = "\n".join(parts).strip()
                    last_result = result

                    result = _extract_sql(result)

                    stop_reason = str(getattr(resp, "stop_reason", ""))
                    if stop_reason != "max_tokens":
                        return result, None

                else:
                    return "", f"Unsupported provider: {provider}"

                next_budget = current_max_tokens * 2
                if next_budget > max_ceiling:
                    return result, "LLM output hit the max_tokens ceiling after retries"
                current_max_tokens = next_budget
                continue

            except Exception as exc:
                response = getattr(exc, "response", None)
                status = getattr(response, "status_code", None) if response is not None else None
                body = getattr(response, "text", None) if response is not None else None

                if status in (429, 502, 503):
                    wait = 2 ** attempt
                    time.sleep(wait)
                    continue

                return last_result, f"LLM fix failed: {exc} (status={status} body={body})"

        return last_result, "LLM output hit the max_tokens ceiling after retries"

    @staticmethod
    def _build_explanation(stats: Dict[str, Any]) -> str:
        lines = ["### Translation Pipeline\n"]
        for step in stats.get("steps", []):
            lines.append(f"- {step}")
        lines.append(
            f"\n**Chunks:** {stats.get('chunks', 1)}  |  "
            f"**LLM calls:** {stats.get('llm_calls', 0)}  |  "
            f"**Cache hits:** {stats.get('cache_hits', 0)}"
        )
        return "\n".join(lines)
